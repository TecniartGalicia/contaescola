"""
migrar_excel.py — Migración completa a ContaEscola v6  (v3 — definitiva)
Ejecutar DENTRO de la carpeta contaescola_v6:
    pip install openpyxl
    python migrar_excel.py

CORRECCIONES v3:
  1. ano = AÑO DEL ARCHIVO (ejercicio contable), NO de la fecha del movimiento.
     Los nums 220-225 tienen fecha 2026 pero pertenecen al ejercicio 2025 porque
     están en FUNCIONAMENTO_2025.xlsx — son pagos de 2025 hechos en enero 2026.

  2. Los archivos PARTIDAS solo hacen UPDATE (alumno_neae, xustifica).
     No insertan movimientos nuevos — ya están todos en los archivos principales.

  3. Check de duplicados en paso 6: (area, num, ano_archivo).
"""
import openpyxl, sqlite3, os, re, sys
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'contaescola.db')

# ano_archivo = ejercicio contable al que pertenece el archivo
ARCHIVOS_DIARIO = [
    ('FUNCIONAMIENTO_2026.xlsx', 'func', 2026, '2025-2026'),
    ('COMEDOR_2026.xlsx',        'com',  2026, '2025-2026'),
    ('FUNCIONAMENTO_2025.xlsx',  'func', 2025, '2024-2025'),
    ('COMEDOR_2025.xlsx',        'com',  2025, '2024-2025'),
]
ARCHIVOS_PARTIDAS = [
    ('PARTIDAS__para_funcionamento_2025.xlsx', 2025),
    ('PARTIDAS_para_funcionamento_2026.xlsx',  2026),
]

# ── CÓDIGOS CONTABLES ─────────────────────────────────────────
CODIGOS_TABLA = [
    ('01',  'Arredamientos',                        1),
    ('02',  'Reparacións',                           2),
    ('03',  'Material de oficina',                   3),
    ('04',  'Suministracións',                       4),
    ('04.1','Gasoleo',                               5),
    ('04.2','Gas',                                   6),
    ('04.3','Biomasa/Pellets',                       7),
    ('04.4','Electricidade',                         8),
    ('04.5','Auga',                                  9),
    ('04.6','Outras Subministracións',              10),
    ('05',  'Comunicacións',                        11),
    ('06',  'Transporte',                           12),
    ('07',  'Traballos realizados outras empresas', 13),
    ('08',  'Prima de seguro',                      14),
    ('09',  'Tributos',                             15),
    ('10',  'Axudas de custo e locomoción',         16),
    ('11',  'Gastos diversos',                      17),
    ('12',  'Mobiliario e utensilios inventariables',18),
    ('13',  'Outros o materiais inventariables',    19),
    ('14',  'Comedores escolares',                  20),
    ('A',   'Total dotación Consellería',           21),
    ('A1',  'Gastos de funcionamento',              22),
    ('A2',  'Fondo solidario libros',               23),
    ('A3',  'Dotación Biblioteca',                  24),
    ('A4',  'Fondos plan mellora',                  25),
    ('A5',  'Fondos innovación',                    26),
    ('A6',  'Dotación comedores escolares',         27),
    ('A7',  'Fondos P Lingüística',                 28),
    ('A8',  'Otros',                                29),
    ('B',   'Dotacións UE-Outros organismos',       30),
    ('C',   'Dotacións legados e subvencións',      31),
    ('D',   'Ingreso convenios',                    32),
    ('E',   'Alleamento de bens',                   33),
    ('F',   'Ingresos vendas',                      34),
    ('G',   'Ingresos uso instalacións',            35),
    ('H',   'SALDO ANTERIOR',                       36),
    ('I',   'Venda fotocopias',                     37),
    ('J',   'Achegas comedor particulares',         38),
    ('K',   'Xuros das contas bancarias',           39),
    ('L',   'Outros ingresos',                      40),
]
COD_DESC = {c: d for c, d, _ in CODIGOS_TABLA}

# ── ALUMNOS NEAE ──────────────────────────────────────────────
ALUMNOS_NEAE = [
    ('CANDELA',     400.0,  '2025-2026'),
    ('CHRISTOPHER', 400.0,  '2025-2026'),
    ('ROI',         1313.0, '2025-2026'),
    ('SAMARA',      2226.0, '2025-2026'),
    ('MATHIAS',     400.0,  '2025-2026'),
]

# ── PARTIDAS CON IMPORTES ─────────────────────────────────────
PARTIDAS_DEF = [
    ('BECAS NEAE',                                       '2024-2025', 4739.0),
    ('CENTROS PIE',                                      '2024-2025', 3000.0),
    ('PLAMBE-FONDOS',                                    '2024-2025', 2575.23),
    ('PLAMBE-MOBILIARIO',                                '2024-2025', 1600.0),
    ('LINGUAS EXTRANXEIRAS',                             '2024-2025', 1055.0),
    ('FOMENTO DO USO DO GALEGO',                         '2024-2025', 367.0),
    ('XESTION ATENCION BIBLIOTECA',                      '2024-2025', 164.2),
    ('X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA','2024-2025', 175.0),
    ('DESENVOLVEMENTO PROGRAMA EDUCACIÓN DIXITAL',       '2024-2025', 175.0),
    ('PROGRAMA ÉXITO ESCOLAR-IMPULSO DAS MATEMÁTICAS',   '2024-2025', 825.0),
    ('BECAS NEAE',                                       '2025-2026', 4739.0),
    ('CENTROS PIE',                                      '2025-2026', 900.0),
    ('PLAMBE-FONDOS',                                    '2025-2026', 0.0),
    ('PLAMBE-MOBILIARIO',                                '2025-2026', 0.0),
    ('LINGUAS EXTRANXEIRAS',                             '2025-2026', 553.98),
    ('FOMENTO DO USO DO GALEGO',                         '2025-2026', 0.0),
    ('XESTION ATENCION BIBLIOTECA',                      '2025-2026', 164.2),
    ('X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA','2025-2026', 0.0),
    ('DESENVOLVEMENTO PROGRAMA EDUCACIÓN DIXITAL',       '2025-2026', 0.0),
    ('PROGRAMA ÉXITO ESCOLAR-IMPULSO DAS MATEMÁTICAS',   '2025-2026', 825.0),
]

# ── NORMALIZACIÓN XUSTIFICA ───────────────────────────────────
NORM_XUST = {
    'PLAMBE FONDOS':                   'PLAMBE-FONDOS',
    'PLAMBE MOBILIARIO':               'PLAMBE-MOBILIARIO',
    'BECA NEAE SAMARA':                'BECAS NEAE',
    'BECA NEAE ROI':                   'BECAS NEAE',
    'BECA NEAE CANDELA':               'BECAS NEAE',
    'BECA NEAE MATHIAS':               'BECAS NEAE',
    'BECA NEAE CHRISTOPHER':           'BECAS NEAE',
    'PIE':                             'CENTROS PIE',
    'EDUCACION DIXITAL':               'DESENVOLVEMENTO PROGRAMA EDUCACIÓN DIXITAL',
    'COMPETENCIA LIGÜISTICA':          'X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA',
    'PROGRAMA COMPETENCIA LINGUÍSTICA':'X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA',
    'X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA':
                                       'X.T. PROGRAMA CERTIFICACIÓN COMPETENCIA LINGÜÍSTICA',
    'FONDO LIBROS':                    'FONDO SOLIDARIO LIBROS EDIXGAL',
    'ANDAINA MONFORTE':                'ACTIVIDADES EXTRAESCOLARES',
    'OBRA COCIÑA':                     'OBRA COCIÑA',
    'MATERIAL':                        '',
    'REPROGRAFIA':                     '',
    'ALARMA':                          '',
}
XUST_ALUMNO = {
    'BECA NEAE SAMARA':       'SAMARA',
    'BECA NEAE ROI':          'ROI',
    'BECA NEAE CANDELA':      'CANDELA',
    'BECA NEAE MATHIAS':      'MATHIAS',
    'BECA NEAE CHRISTOPHER':  'CHRISTOPHER',
    'BECAS NEAE MATHIAS (CURSO 24/25)': 'MATHIAS',
}

# ── NORMALIZACIÓN CLIENTES ────────────────────────────────────
CANONICO = {
    'GRENKE RENT, S.L.U.':                 'GRENKE RENT S.L.U.',
    '\tGRENKE RENT, S.L.U.':               'GRENKE RENT S.L.U.',
    'MERCACHINA':                           'MERCA CHINA',
    'MIDETERGENTE':                         'MI DETERGENTE',
    'NEX CENTER BAZAR MODA':                'NEW CENTRE BAZAR MODA',
    'NEW CENTER':                           'NEW CENTRE BAZAR MODA',
    'REPSOL BUTANO, S.A.':                  'REPSOL BUTANO S.A.',
    'REPSOL BUTANO':                        'REPSOL BUTANO S.A.',
    'GRAN BAZAR LIYONG':                    'BAZAR LIYONG S.L.',
    'RESERVA LUCUS AVENTUR':                'LUCUS AVENTUR',
    'RESERVA AC. LUCUS AVENTURE':           'LUCUS AVENTUR',
    '2º PAGO AC. LUCUS AVENTUR':            'LUCUS AVENTUR',
    '2º ABONO LUCUS AVENTUR 3º/4º':        'LUCUS AVENTUR',
    'EXCUR. 5º LUCUS AVENTUR':              'LUCUS AVENTUR',
    'ITINERANCIAS 1º TRIMESTRE NATI':       'ITINERANCIAS NATI',
    'ITINERANCIAS XANEIRO NATI':            'ITINERANCIAS NATI',
    'ITINERANCIAS FEBREIRO NATI':           'ITINERANCIAS NATI',
    'ITINERANCIAS ABRIL NATI':              'ITINERANCIAS NATI',
    'IITINERANCIAS MAIO NATI':              'ITINERANCIAS NATI',
    'ITINERANCIAS XUÑO NATI':               'ITINERANCIAS NATI',
    'ITINERANCIAS NATI MARZO 2025':         'ITINERANCIAS NATI',
    'ITINERANCIAS MAIO CRISTINA':           'ITINERANCIAS CRISTINA',
    'ITINERANCIAS 1º TRIMESTRE BEGOÑA':     'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS XANEIRO BEGONA':          'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS FEBREIRO BEGONA':         'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS ABRIL BEGONA':            'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS MAIO 2025 BEGONA':        'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS XUÑO BEGOÑA':             'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS BEGONA MARZO 2025':       'ITINERANCIAS BEGOÑA',
    'ITINERANCIAS XANEIRO CRISTINA':        'ITINERANCIAS CRISTINA',
    'ITINERANCIAS FEBREIRO CRISTINA':       'ITINERANCIAS CRISTINA',
    'ITINERANCIAS ABRIL CRISTINA':          'ITINERANCIAS CRISTINA',
    'ITINERANCIAS CRISTINA MARZO 2025':     'ITINERANCIAS CRISTINA',
    'ITINERANCIAS XANEIRO PAMELA':          'ITINERANCIAS PAMELA',
    'ITINERANCIAS FEBREIRO PAMELA':         'ITINERANCIAS PAMELA',
    'DIETA 29/05':                          'DIETAS PROFESORADO',
    'DIETA 29/06':                          'DIETAS PROFESORADO',
    'DIETAS 26,27,28,29/06':               'DIETAS PROFESORADO',
    'DIETAS CONSERXERÍA':                   'DIETAS PROFESORADO',
    'VIAXE A LUGO XEFATURA E SECRETARÍA':         'VIAXES E DESPRAZAMENTOS',
    'VIAXE A MONFORTE MARÍA ISABEL MAZAIRA':      'VIAXES E DESPRAZAMENTOS',
    'VIAXE NATI SANTIAGO':                        'VIAXES E DESPRAZAMENTOS',
    'VIAXE-COMIDA DIRECTOR SANTIAGO 20/0 3/25':   'VIAXES E DESPRAZAMENTOS',
    'DESPRAZAMENTOS AO CENTRO POR OBRAS DIRECTOR':'VIAXES E DESPRAZAMENTOS',
    'PARKING LUGO':                         'VIAXES E DESPRAZAMENTOS',
    'AUTOBUS 1º CICLO':                     'AUTOBUSES EXCURSIONS',
    'AUTOBUS E. INFANTIL':                  'AUTOBUSES EXCURSIONS',
    'AUTOBÚS 3º/4º (BILLETES)':            'AUTOBUSES EXCURSIONS',
    'AUTOBÚS 3º/4º (MOEDAS)':             'AUTOBUSES EXCURSIONS',
    'PSICÓLOGA CRISTINA VIDAL':             'CRISTINA VIDAL',
    'ANA RODRIGUEZ (XUSTIFICADO 2025)':     'ANA RODRÍGUEZ',
    'ANA RODRÍGUEZ (XUSTIFICADO 2025)':    'ANA RODRÍGUEZ',
    'ANA RODRIGUEZ':                        'ANA RODRÍGUEZ',
    'JOSE LUIS RODRIGUEZ BLANCO':           'JOSÉ LUIS RODRÍGUEZ BLANCO',
    'ANA BELEN TRILLO LODEIRO':             'ANA BELEN LODEIRO TRILLO',
    'LIBRERÍA MAGO':                        'LIBRERIA MAGO',
    'ELSUNDO':                              'ELSUNO',
    'AMAZON EU':                            'AMAZON',
    'X.T- CENTROS PIE 2024':               'DEPARTAMENTO TERRITORIAL',
    'X.T- CENTROS PIE 2025':               'DEPARTAMENTO TERRITORIAL',
    'X.T.PROGRAMA CERTIFICACION COMPETENCIA LINGÜÍSTICA':'DEPARTAMENTO TERRITORIAL',
    'X.T. APRENDIZAXE LINGUAS EXTRAXEIRAS 2024':'DEPARTAMENTO TERRITORIAL',
    'X.T. APRENDIZAXE LINGUAS EXTRAXEIRAS 2026':'DEPARTAMENTO TERRITORIAL',
    'X.T. DESENVOLVEMENTO PROG EDUCACION DIXITAL':'DEPARTAMENTO TERRITORIAL',
    'X.T.PLAMBE MOBILIARIO':               'DEPARTAMENTO TERRITORIAL',
    'X.T. XESTION ATENCION BIBLIOTECA ESCOLAR':'DEPARTAMENTO TERRITORIAL',
    'X.T. FOMENTO USO DO GALEGO':          'DEPARTAMENTO TERRITORIAL',
    'X.T. PLAMBE FONDOS':                  'DEPARTAMENTO TERRITORIAL',
    'BECAS S G TESORO':                    'BECAS S.G. TESORO',
    'CENTROS PIE 2025':                    'DEPARTAMENTO TERRITORIAL',
    'X.T CENTROS PIE 2025':               'DEPARTAMENTO TERRITORIAL',
    'DECLARACION IMPUESTOS':               'DECLARACIÓN IMPUESTOS',
    'PROGRAMA ÉXITO ESCOLAR-IMPULSO DAS MATEMÁTICAS':'DEPARTAMENTO TERRITORIAL',
    'FONDO SOLIDARIO DE LIBROS EDIXGAL':   'FONDO SOLIDARIO DE LIBROS EDIXGAL',
}

# ── HELPERS ───────────────────────────────────────────────────
def clean(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s).strip().upper())

def norm_cl(raw):
    c = clean(raw); return CANONICO.get(c, c)

def norm_xust(raw):
    c = clean(raw); return NORM_XUST.get(c, c)

def alumno_xust(raw):
    c = clean(raw); return XUST_ALUMNO.get(c, '')

def parse_fecha(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
            try: return datetime.strptime(v.strip(), fmt).strftime('%Y-%m-%d')
            except: pass
    return None

def trim(fecha):
    try:
        m = datetime.strptime(fecha, '%Y-%m-%d').month
        if m<=3: return '1º TRIMESTRE'
        if m<=6: return '2º TRIMESTRE'
        if m<=9: return '3º TRIMESTRE'
        return '4º TRIMESTRE'
    except: return ''

def fmt(n): return f"{n:,.2f}".replace(',','X').replace('.',',').replace('X','.')

# ── MAIN ──────────────────────────────────────────────────────
def main():
    for fname, *_ in ARCHIVOS_DIARIO + ARCHIVOS_PARTIDAS:
        if not os.path.exists(fname):
            print(f"❌ Arquivo non atopado: {fname}"); sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"❌ BD non atopada: {DB_PATH}")
        print("   Arranca a app polo menos unha vez antes de migrar.")
        sys.exit(1)

    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA foreign_keys = ON")

    print("=" * 60)
    print("  MIGRACIÓN COMPLETA v3 — ContaEscola")
    print("=" * 60)

    # ── 1. Códigos ─────────────────────────────────────────────
    print("\n[1/8] Actualizando códigos contables...")
    con.execute("DELETE FROM codigos")
    for cod, desc, orden in CODIGOS_TABLA:
        con.execute("INSERT INTO codigos (codigo,descripcion,activo,orden) VALUES (?,?,1,?)",
                    (cod, desc, orden))
    con.commit()
    print(f"   ✅ {len(CODIGOS_TABLA)} códigos")

    # ── 2. Anos e cursos ───────────────────────────────────────
    print("\n[2/8] Preparando anos e cursos...")
    for a in [2025, 2026]:
        con.execute("INSERT OR IGNORE INTO anos VALUES (?)", (a,))
    for nome in ['2024-2025', '2025-2026']:
        con.execute("INSERT OR IGNORE INTO cursos (nome) VALUES (?)", (nome,))
    con.commit()
    cur_ids = {r[0]:r[1] for r in con.execute("SELECT nome,id FROM cursos").fetchall()}

    # ── 3. Alumnos NEAE ────────────────────────────────────────
    print("\n[3/8] Alumnos NEAE...")
    al_n = 0
    for nome, beca, curso_nome in ALUMNOS_NEAE:
        cur_id = cur_ids.get(curso_nome)
        ex = con.execute("SELECT id FROM alumnos_neae WHERE nome=?", (nome,)).fetchone()
        if ex:
            con.execute("UPDATE alumnos_neae SET importe_beca=?, curso_id=? WHERE nome=?",
                        (beca, cur_id, nome))
        else:
            con.execute("""INSERT INTO alumnos_neae
                (nome,curso_id,curso_ingreso,importe_beca,notas) VALUES (?,?,?,?,?)""",
                (nome, cur_id, '', beca, 'Importado automaticamente'))
            al_n += 1
        print(f"   ✅ {nome:<14} beca: {fmt(beca)} €")
    con.commit()

    # ── 4. Partidas ────────────────────────────────────────────
    print("\n[4/8] Partidas finalistas con importes...")
    p_n = p_a = 0
    for nome, curso_nome, imp in PARTIDAS_DEF:
        cur_id = cur_ids.get(curso_nome)
        if not cur_id: continue
        ex = con.execute("SELECT id FROM partidas_config WHERE curso_id=? AND nome=?",
                         (cur_id, nome)).fetchone()
        if ex:
            con.execute("UPDATE partidas_config SET importe_asignado=? WHERE id=?", (imp, ex[0]))
            p_a += 1
        else:
            con.execute("INSERT INTO partidas_config (curso_id,nome,importe_asignado,notas) VALUES (?,?,?,?)",
                        (cur_id, nome, imp, 'Importada automaticamente'))
            p_n += 1
    con.commit()
    print(f"   ✅ {p_n} novas · {p_a} actualizadas")

    # ── 5. Clientes ────────────────────────────────────────────
    print("\n[5/8] Clientes normalizados...")
    todos_canon = set()
    for fname, *_ in ARCHIVOS_DIARIO + ARCHIVOS_PARTIDAS:
        wb = openpyxl.load_workbook(fname, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            c = clean(row[2])
            if not c or c == 'SALDO ANTERIOR': continue
            if float(row[3] or row[4] or 0) <= 0: continue
            n = norm_cl(c)
            if n: todos_canon.add(n)
    cl_n = 0
    for nome in sorted(todos_canon):
        if not nome: continue
        if not con.execute("SELECT id FROM clientes WHERE nome=?", (nome,)).fetchone():
            con.execute("INSERT INTO clientes (nome,tipo,nif,direccion,telefono,email,notas) VALUES (?,?,?,?,?,?,?)",
                        (nome,'proveedor','','','','','Importado automaticamente'))
            cl_n += 1
    con.commit()
    cl_map = {r[0]:r[1] for r in con.execute("SELECT nome,id FROM clientes").fetchall()}
    print(f"   ✅ {cl_n} novos · {len(cl_map)} total")

    # ── 6. DIARIO PRINCIPAL ────────────────────────────────────
    # ★ CORRECCIÓN 1: ano = ano_archivo (ejercicio contable del archivo)
    # ★ CORRECCIÓN 2: check duplicados por (area, num, ano_archivo)
    print("\n[6/8] Importando diario principal...")
    total_ins = total_skip = total_err = 0

    for fname, area, ano_archivo, curso_nome in ARCHIVOS_DIARIO:
        cur_id  = cur_ids.get(curso_nome)
        es_func = area == 'func'
        col_per = 8 if es_func else 7
        col_xus = 10 if es_func else 9
        col_nom = 11

        wb = openpyxl.load_workbook(fname, data_only=True)
        ws = wb[wb.sheetnames[0]]
        ins_arch = skip_arch = 0
        saldo = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            concepto = clean(row[2])
            debe = row[3]; haber = row[4]
            ig   = clean(row[5])
            if not concepto: continue

            if concepto == 'SALDO ANTERIOR':
                saldo = float(haber or 0)
                con.execute("INSERT OR REPLACE INTO saldos VALUES (?,?,?)", (ano_archivo, area, saldo))
                continue

            importe = float(debe or haber or 0)
            if importe <= 0: continue
            fecha = parse_fecha(row[1])
            if not fecha: continue

            try: num_int = int(row[0] or 0)
            except: num_int = 0

            tipo    = ig if ig in ('G','I') else ('G' if debe else 'I')
            cod_raw = clean(row[6]) if len(row)>6 else ''
            periodo = clean(row[col_per]) if col_per<len(row) else ''
            xust_r  = clean(row[col_xus]) if col_xus<len(row) else ''
            nome_al = clean(row[col_nom]) if col_nom<len(row) else ''

            if es_func:
                codigo = cod_raw if cod_raw not in ('H','') else ''
                cod_d  = COD_DESC.get(codigo, '')
                categ  = ''
            else:
                codigo = ''; cod_d = ''
                NCAT = {'ALIMENTACIÓN':'ALIMENTACION','ALIMENTACION':'ALIMENTACION',
                        'LIMPEZA':'LIMPEZA','COMBUSTIBLE':'COMBUSTIBLE','MANTEMENTO':'MANTEMENTO'}
                categ = NCAT.get(cod_raw, cod_raw or 'OUTROS')

            xust   = norm_xust(xust_r)
            alumno = nome_al or alumno_xust(xust_r)
            cl_id  = cl_map.get(norm_cl(concepto))

            # ★ Check: (area, num, ano_archivo) — el número es único por ejercicio
            dup = con.execute(
                "SELECT id, alumno_neae, xustifica FROM diario WHERE area=? AND num=? AND ano=?",
                (area, num_int, ano_archivo)
            ).fetchone()
            if dup:
                new_al = alumno or dup[1] or ''
                new_x  = xust   or dup[2] or ''
                if new_al != (dup[1] or '') or new_x != (dup[2] or ''):
                    con.execute("UPDATE diario SET alumno_neae=?, xustifica=? WHERE id=?",
                                (new_al, new_x, dup[0]))
                total_skip += 1; skip_arch += 1; continue

            try:
                con.execute("""INSERT INTO diario
                    (area,ano,curso_id,num,data,tipo,importe,concepto,
                     codigo,cod_desc,periodo,notas,categoria,xustifica,alumno_neae,cliente_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (area, ano_archivo, cur_id, num_int, fecha, tipo, importe,
                     'PAGO FACTURA', codigo, cod_d,
                     periodo or trim(fecha), '', categ, xust, alumno, cl_id))
                total_ins += 1; ins_arch += 1
            except Exception as e:
                total_err += 1
                if total_err <= 5:
                    print(f"   ⚠️  {fname} num={num_int}: {e}")

        con.commit()
        s = fmt(saldo)+' €' if saldo else 'non detectado'
        print(f"   ✅ {fname:<40} {ins_arch:>3} insertados · {skip_arch} saltados · saldo: {s}")

    # ── 7. PARTIDAS — SOLO UPDATE, NUNCA INSERT ────────────────
    # ★ CORRECCIÓN 3: los archivos PARTIDAS no añaden movimientos nuevos.
    #   Solo enriquecen alumno_neae y xustifica de movimientos ya insertados.
    print("\n[7/8] Enriquecendo alumno_neae e xustifica desde arquivos PARTIDAS...")
    p_upd = p_nf = 0

    for fname, ano_archivo in ARCHIVOS_PARTIDAS:
        wb = openpyxl.load_workbook(fname, data_only=True)
        ws = wb[wb.sheetnames[0]]
        upd_arch = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            xust_r  = clean(row[10]) if len(row)>10 else ''
            nome_al = clean(row[11]) if len(row)>11 else ''
            if not xust_r and not nome_al: continue  # sin info nueva, saltar

            try: num_int = int(row[0])
            except (TypeError, ValueError): continue
            if num_int == 0: continue

            xust   = norm_xust(xust_r)
            alumno = nome_al or alumno_xust(xust_r)

            # Buscar el registro existente por (area, num, ano_archivo)
            ex = con.execute(
                "SELECT id, alumno_neae, xustifica FROM diario WHERE area='func' AND num=? AND ano=?",
                (num_int, ano_archivo)
            ).fetchone()

            if not ex:
                # Puede estar en el otro año — buscar sin filtrar año
                ex = con.execute(
                    "SELECT id, alumno_neae, xustifica FROM diario WHERE area='func' AND num=?",
                    (num_int,)
                ).fetchone()

            if ex:
                new_al = alumno or ex[1] or ''
                new_x  = xust   or ex[2] or ''
                if new_al != (ex[1] or '') or new_x != (ex[2] or ''):
                    con.execute("UPDATE diario SET alumno_neae=?, xustifica=? WHERE id=?",
                                (new_al, new_x, ex[0]))
                    upd_arch += 1; p_upd += 1
            else:
                p_nf += 1

        con.commit()
        print(f"   ✅ {fname:<45} {upd_arch:>3} actualizados")

    if p_nf > 0:
        print(f"   ℹ️  {p_nf} referencias en PARTIDAS sen movemento correspondente (normal)")

    # ── 8. Normalizar xustificas residuais ────────────────────
    print("\n[8/8] Normalizando xustificas residuais...")
    act = 0
    for old, new in NORM_XUST.items():
        if not new: continue
        r = con.execute("UPDATE diario SET xustifica=? WHERE xustifica=?", (new, old))
        if r.rowcount > 0:
            print(f"   ↺  '{old}' → '{new}' ({r.rowcount})")
            act += r.rowcount
    con.commit()

    # ── RESUMEN FINAL ─────────────────────────────────────────
    resumen = con.execute("""
        SELECT ano, area,
               COUNT(*) as n,
               SUM(CASE WHEN tipo='G' THEN importe ELSE 0 END) as debe,
               SUM(CASE WHEN tipo='I' THEN importe ELSE 0 END) as haber
        FROM diario GROUP BY ano, area ORDER BY ano, area
    """).fetchall()
    con.close()

    print()
    print("=" * 60)
    print("  ✅  MIGRACIÓN COMPLETADA v3")
    print("=" * 60)
    print(f"  Movementos insertados : {total_ins}")
    print(f"  Saltados (xa existían): {total_skip}")
    print(f"  Campos actualizados   : {p_upd}")
    print(f"  Erros                 : {total_err}")
    print()
    print(f"  {'Ano':<6} {'Área':<6} {'Movs':>5}  {'Debe €':>13}  {'Haber €':>13}")
    print("  " + "-"*50)
    for r in resumen:
        print(f"  {r[0]:<6} {r[1]:<6} {r[2]:>5}  {fmt(r[3]):>13}  {fmt(r[4]):>13}")
    print()
    print("  Totais esperados (Funcionamento):")
    print("  2025 func → Debe: 65.985,68 €  Haber: 48.933,17 €")
    print("  2026 func → Debe: 13.009,84 €  Haber:     80,00 €")
    print("=" * 60)

if __name__ == '__main__':
    main()
